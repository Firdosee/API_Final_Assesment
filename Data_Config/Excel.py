import openpyxl


class Excel_Data:

    def getAPIData(self, test_name):
        try:
            data_dict = {}
            book = openpyxl.load_workbook("../API_Data.xlsx")
            sheet = book.active
            for i in range(1, sheet.max_row + 1):  # to get rows
                if sheet.cell(row=i, column=1).value == test_name:

                    for j in range(2, sheet.max_column + 1):  # to get columns
                        data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            return data_dict
        except FileNotFoundError as e:
            print("Exception occurred while performing file operation", e)

    def gettaskData(self, test_name):
        try:
            data_dict = {}
            book = openpyxl.load_workbook("../API_Data.xlsx")
            sheet = book.active
            for i in range(1, sheet.max_row + 1):  # to get rows
                if sheet.cell(row=i, column=1).value == test_name:

                    for j in range(2, sheet.max_column + 1):  # to get columns
                        data_dict[sheet.cell(row=11, column=j).value] = sheet.cell(row=i, column=j).value
            return data_dict
        except FileNotFoundError as e:
            print("Exception occurred while performing file operation", e)
